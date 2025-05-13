
import streamlit as st
import pandas as pd
import numpy as np
import json
import os
from datetime import datetime
from io import StringIO, BytesIO
import copy
import plotly.io as pio
import matplotlib
import matplotlib.pyplot as plt
import base64
import time
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import matplotlib.pyplot as plt
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet
from PIL import Image as PILImage
from utils import calculate_peso_maximo_efectivo
matplotlib.use('Agg')

from utils import load_csv_with_fallback, clasificar_base_refinada
from calculations import sugerencias_final_con_fak, check_cumulative_weights, calculate_final_values
from manual_calculation import manual_assignment
from automatic_calculation import automatic_assignment
from visualizations import print_final_summary, plot_main_deck, plot_lower_decks
from data_models import FlightData, AircraftData, CalculationState, FinalResults

class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NumpyEncoder, self).default(obj)

def get_unique_filename(base_path, extension):
    """Genera un nombre de archivo único añadiendo un sufijo numérico si es necesario."""
    base_name = os.path.splitext(base_path)[0]
    counter = 1
    new_path = base_path
    while os.path.exists(new_path):
        new_path = f"{base_name}_{counter}.{extension}"
        counter += 1
    return new_path
class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NumpyEncoder, self).default(obj)

def weight_balance_calculation():
    try:
        from utils import load_csv_with_fallback, clasificar_base_refinada
        from calculations import sugerencias_final_con_fak, check_cumulative_weights, calculate_final_values
        from manual_calculation import manual_assignment
        from automatic_calculation import automatic_assignment
        from visualizations import print_final_summary, plot_main_deck, plot_lower_decks
        from data_models import FlightData, AircraftData, CalculationState, FinalResults
    except ImportError as e:
        st.error(f"Error al importar módulos: {str(e)}. Verifique que todos los archivos necesarios estén en el directorio correcto.")
        st.stop()

    st.title("Sistema de Cálculo de Peso y Balance")

    if "manifiesto_manual" not in st.session_state:
        st.session_state.manifiesto_manual = pd.DataFrame({
            "Contour": [""],
            "Number ULD": [""],
            "ULD Final Destination": [""],
            "Weight (KGS)": [0.0],
            "Pieces": [0],
            "Notes": [""]
        })

    st.markdown('<div id="carga_datos_iniciales_section"></div>', unsafe_allow_html=True)
    st.subheader("Carga de Datos Iniciales")
    
    default_flight_data = {
        "operador": "TAMPA CARGO",
        "numero_vuelo": "TPA",
        "matricula": "",
        "fecha_vuelo": "",
        "hora_vuelo": "",
        "ruta_vuelo": "",
        "revision": "",
        "destino_inicial": "",
        "fuel_kg": 0.0,
        "trip_fuel": 0.0,
        "taxi_fuel": 0.0,
        "tipo_carga": "Simétrico",
        "takeoff_runway": "",
        "rwy_condition": "Dry",
        "flaps_conf": "1+F",
        "temperature": 0.0,
        "air_condition": "Off",
        "anti_ice": "Off",
        "qnh": 1013.0,
        "performance_tow": 0.0,
        "performance_lw": 0.0,
        "passengers_cockpit": 0,
        "passengers_supernumerary": 0,
        "ballast_fuel": 0.0
    }
    
    default_calc_state = {
        "df": None,
        "posiciones_usadas": set(),
        "rotaciones": {},
        "bow": 0.0,
        "bow_moment_x": 0.0,
        "bow_moment_y": 0.0,
        "moment_x_fuel_tow": 0.0,
        "moment_y_fuel_tow": 0.0,
        "moment_x_fuel_lw": 0.0,
        "moment_y_fuel_lw": 0.0,
        "passengers_cockpit_total_weight": 0.0,
        "passengers_cockpit_total_moment_x": 0.0,
        "passengers_supernumerary_total_weight": 0.0,
        "passengers_supernumerary_total_moment_x": 0.0,
        "fuel_distribution": {
            "Outer Tank LH": 0.0,
            "Outer Tank RH": 0.0,
            "Inner Tank LH": 0.0,
            "Inner Tank RH": 0.0,
            "Center Tank": 0.0,
            "Trim Tank": 0.0
        },
        "fuel_mode": "Automático"
    }

    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = script_dir
    aircraft_db_path = os.path.join(base_dir, "General_aircraft_database.csv")

    if not os.path.exists(aircraft_db_path):
        st.error(f"No se encontró el archivo en: {aircraft_db_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        st.stop()
    aircraft_db = pd.read_csv(aircraft_db_path, sep=";", decimal=",")

    # Handle JSON import
    if "json_imported" in st.session_state and st.session_state.json_imported:
        try:
            if not hasattr(st.session_state.json_imported, 'read'):
                st.error("El archivo JSON no es válido o no se cargó correctamente.")
                return
            
            json_content = st.session_state.json_imported.read()
            if not json_content:
                st.error("El archivo JSON está vacío.")
                return
            
            json_data = json.loads(json_content)
            if json_data is None:
                st.error("El archivo JSON no contiene datos válidos.")
                return
            
            flight_info = json_data.get("flight_info", {})
            calculated_values = json_data.get("calculated_values", {})
            passengers = json_data.get("passengers", {})
            takeoff_conditions = json_data.get("takeoff_conditions", {})
            manifest_data = json_data.get("manifest_data", [])
            posiciones_usadas = set(json_data.get("posiciones_usadas", []))
            rotaciones = json_data.get("rotaciones", {})
            tipo_carga = json_data.get("tipo_carga", "Simétrico").capitalize()
            fuel_distribution = calculated_values.get("fuel_distribution", default_calc_state["fuel_distribution"])
            fuel_mode = calculated_values.get("fuel_mode", "Automático")

            default_flight_data.update({
                "operador": flight_info.get("operador", default_flight_data["operador"]),
                "numero_vuelo": flight_info.get("numero_vuelo", default_flight_data["numero_vuelo"]),
                "matricula": flight_info.get("matricula", default_flight_data["matricula"]),
                "fecha_vuelo": flight_info.get("fecha_vuelo", default_flight_data["fecha_vuelo"]),
                "hora_vuelo": flight_info.get("hora_vuelo", default_flight_data["hora_vuelo"]),
                "ruta_vuelo": flight_info.get("ruta_vuelo", default_flight_data["ruta_vuelo"]),
                "revision": flight_info.get("revision", default_flight_data["revision"]),
                "destino_inicial": flight_info.get("destino_inicial", default_flight_data["destino_inicial"]),
                "fuel_kg": calculated_values.get("fuel_kg", default_flight_data["fuel_kg"]),
                "trip_fuel": calculated_values.get("trip_fuel", default_flight_data["trip_fuel"]),
                "taxi_fuel": calculated_values.get("taxi_fuel", default_flight_data["taxi_fuel"]),
                "tipo_carga": tipo_carga,
                "takeoff_runway": takeoff_conditions.get("runway", default_flight_data["takeoff_runway"]),
                "rwy_condition": takeoff_conditions.get("rwy_condition", default_flight_data["rwy_condition"]),
                "flaps_conf": takeoff_conditions.get("flaps_conf", default_flight_data["flaps_conf"]),
                "temperature": takeoff_conditions.get("temperature", default_flight_data["temperature"]),
                "air_condition": takeoff_conditions.get("air_condition", default_flight_data["air_condition"]),
                "anti_ice": takeoff_conditions.get("anti_ice", default_flight_data["anti_ice"]),
                "qnh": takeoff_conditions.get("qnh", default_flight_data["qnh"]),
                "performance_tow": takeoff_conditions.get("performance_tow", default_flight_data["performance_tow"]),
                "performance_lw": takeoff_conditions.get("performance_lw", default_flight_data["performance_lw"]),
                "passengers_cockpit": passengers.get("cockpit", default_flight_data["passengers_cockpit"]),
                "passengers_supernumerary": passengers.get("supernumerary", default_flight_data["passengers_supernumerary"]),
                "ballast_fuel": calculated_values.get("ballast_fuel", default_flight_data["ballast_fuel"])
            })

            default_calc_state.update({
                "df": pd.DataFrame(manifest_data) if manifest_data else None,
                "posiciones_usadas": posiciones_usadas,
                "rotaciones": rotaciones,
                "bow": calculated_values.get("bow", default_calc_state["bow"]),
                "bow_moment_x": calculated_values.get("bow_moment_x", default_calc_state["bow_moment_x"]),
                "bow_moment_y": calculated_values.get("bow_moment_y", default_calc_state["bow_moment_y"]),
                "moment_x_fuel_tow": calculated_values.get("moment_x_fuel_tow", default_calc_state["moment_x_fuel_tow"]),
                "moment_y_fuel_tow": calculated_values.get("moment_y_fuel_tow", default_calc_state["moment_y_fuel_tow"]),
                "moment_x_fuel_lw": calculated_values.get("moment_x_fuel_lw", default_calc_state["moment_x_fuel_lw"]),
                "moment_y_fuel_lw": calculated_values.get("moment_y_fuel_lw", default_calc_state["moment_y_fuel_lw"]),
                "passengers_cockpit_total_weight": passengers.get("cockpit_weight", default_calc_state["passengers_cockpit_total_weight"]),
                "passengers_cockpit_total_moment_x": passengers.get("cockpit_moment_x", default_calc_state["passengers_cockpit_total_moment_x"]),
                "passengers_supernumerary_total_weight": passengers.get("supernumerary_weight", default_calc_state["passengers_supernumerary_total_weight"]),
                "passengers_supernumerary_total_moment_x": passengers.get("supernumerary_moment_x", default_calc_state["passengers_supernumerary_total_moment_x"]),
                "fuel_distribution": fuel_distribution,
                "fuel_mode": fuel_mode
            })

            matricula_from_json = flight_info.get("matricula", "")
            if matricula_from_json and matricula_from_json in aircraft_db["Tail"].tolist():
                st.session_state.selected_tail = matricula_from_json
            else:
                st.session_state.selected_tail = aircraft_db["Tail"].iloc[0]

            if "calculation_state" not in st.session_state:
                st.session_state.calculation_state = CalculationState(**default_calc_state)
            else:
                st.session_state.calculation_state.df = default_calc_state["df"].copy() if default_calc_state["df"] is not None else None
                st.session_state.calculation_state.posiciones_usadas = default_calc_state["posiciones_usadas"].copy()
                st.session_state.calculation_state.rotaciones = default_calc_state["rotaciones"].copy()
                st.session_state.calculation_state.bow = default_calc_state["bow"]
                st.session_state.calculation_state.bow_moment_x = default_calc_state["bow_moment_x"]
                st.session_state.calculation_state.bow_moment_y = default_calc_state["bow_moment_y"]
                st.session_state.calculation_state.moment_x_fuel_tow = default_calc_state["moment_x_fuel_tow"]
                st.session_state.calculation_state.moment_y_fuel_tow = default_calc_state["moment_y_fuel_tow"]
                st.session_state.calculation_state.moment_x_fuel_lw = default_calc_state["moment_x_fuel_lw"]
                st.session_state.calculation_state.moment_y_fuel_lw = default_calc_state["moment_y_fuel_lw"]
                st.session_state.calculation_state.passengers_cockpit_total_weight = default_calc_state["passengers_cockpit_total_weight"]
                st.session_state.calculation_state.passengers_cockpit_total_moment_x = default_calc_state["passengers_cockpit_total_moment_x"]
                st.session_state.calculation_state.passengers_supernumerary_total_weight = default_calc_state["passengers_supernumerary_total_weight"]
                st.session_state.calculation_state.passengers_supernumerary_total_moment_x = default_calc_state["passengers_supernumerary_total_moment_x"]
                st.session_state.calculation_state.fuel_distribution = default_calc_state["fuel_distribution"].copy()
                st.session_state.calculation_state.fuel_mode = default_calc_state["fuel_mode"]

            if default_calc_state["df"] is not None:
                st.session_state.manifiesto_manual = default_calc_state["df"].copy()
            else:
                st.session_state.manifiesto_manual = pd.DataFrame({
                    "Contour": [""],
                    "Number ULD": [""],
                    "ULD Final Destination": [""],
                    "Weight (KGS)": [0.0],
                    "Pieces": [0],
                    "Notes": [""]
                })

            st.session_state.normal_fuel = float(calculated_values.get("fuel_kg", 0.0) - calculated_values.get("ballast_fuel", 0.0))
            if calculated_values.get("ballast_fuel", 0.0) > 0 and fuel_distribution.get("Trim Tank", 0.0) > 0:
                st.session_state.computed_ballast_fuel = float(calculated_values.get("ballast_fuel", 0.0))
                st.session_state.trapped_fuel = True
            else:
                st.session_state.computed_ballast_fuel = 0.0
                st.session_state.trapped_fuel = False
            st.session_state.trip_fuel = float(calculated_values.get("trip_fuel", 0.0))
            st.session_state.taxi_fuel = float(calculated_values.get("taxi_fuel", 0.0))
            st.session_state.fuel_mode = fuel_mode
            st.session_state.tipo_carga = tipo_carga
            st.session_state.destino_inicial = flight_info.get("destino_inicial", "")
            st.session_state.takeoff_runway = takeoff_conditions.get("runway", "")
            st.session_state.rwy_condition = takeoff_conditions.get("rwy_condition", "Dry")
            st.session_state.flaps_conf = takeoff_conditions.get("flaps_conf", "1+F")
            st.session_state.temperature = float(takeoff_conditions.get("temperature", 0.0))
            st.session_state.air_condition = takeoff_conditions.get("air_condition", "Off")
            st.session_state.anti_ice = takeoff_conditions.get("anti_ice", "Off")
            st.session_state.qnh = float(takeoff_conditions.get("qnh", 1013.0))
            st.session_state.performance_tow = float(takeoff_conditions.get("performance_tow", 0.0))
            st.session_state.performance_lw = float(takeoff_conditions.get("performance_lw", 0.0))
            st.session_state.passengers_cockpit = int(passengers.get("cockpit", 0))
            st.session_state.passengers_supernumerary = int(passengers.get("supernumerary", 0))
            st.session_state.operador_manual = flight_info.get("operador", default_flight_data["operador"])
            st.session_state.numero_vuelo_manual = flight_info.get("numero_vuelo", default_flight_data["numero_vuelo"])
            st.session_state.matricula_manual = flight_info.get("matricula", default_flight_data["matricula"])
            st.session_state.fecha_vuelo_manual = flight_info.get("fecha_vuelo", default_flight_data["fecha_vuelo"])
            st.session_state.hora_vuelo_manual = flight_info.get("hora_vuelo", default_flight_data["hora_vuelo"])
            st.session_state.ruta_vuelo_manual = flight_info.get("ruta_vuelo", default_flight_data["ruta_vuelo"])
            st.session_state.revision_manual = flight_info.get("revision", default_flight_data["revision"])

            for tank in default_calc_state["fuel_distribution"]:
                st.session_state[f"tank_{tank}"] = float(fuel_distribution.get(tank, 0.0))

            st.session_state.json_imported = None

            if default_calc_state["df"] is not None:
                st.write("Manifiesto importado desde JSON:", default_calc_state["df"])

            st.success("JSON cargado correctamente. Todos los campos han sido prellenados con los datos del archivo. Puede modificarlos y continuar con los cálculos.")
        except json.JSONDecodeError as e:
            st.error(f"Error al decodificar el JSON: {str(e)}. Verifique que el archivo sea un JSON válido.")
            return
        except Exception as e:
            st.error(f"Error al cargar el JSON: {str(e)}")
            return

    if not os.path.exists(aircraft_db_path):
        st.error(f"No se encontró el archivo en: {aircraft_db_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        return
    aircraft_db = pd.read_csv(aircraft_db_path, sep=";", decimal=",")
    
    if "selected_tail" not in st.session_state:
        initial_tail = default_flight_data["matricula"] if default_flight_data["matricula"] in aircraft_db["Tail"].tolist() else aircraft_db["Tail"].tolist()[0]
        st.session_state.selected_tail = initial_tail

    tail_options = aircraft_db["Tail"].tolist()
    tail_index = tail_options.index(st.session_state.selected_tail) if st.session_state.selected_tail in tail_options else 0

    tail = st.selectbox(
        "Seleccione la matrícula de la aeronave",
        tail_options,
        index=tail_index,
        key="tail_selectbox"
    )

    if tail != st.session_state.selected_tail:
        st.session_state.selected_tail = tail
    
    aircraft_folder = os.path.normpath(os.path.join(base_dir, tail))
    if not os.path.exists(aircraft_folder):
        st.error(f"La carpeta {aircraft_folder} no existe.")
        return

    basic_data_path = os.path.join(aircraft_folder, "basic_data.csv")
    if not os.path.exists(basic_data_path):
        st.error(f"No se encontró el archivo en: {basic_data_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        return
    basic_data = pd.read_csv(basic_data_path, sep=";", decimal=",")

    st.markdown('<div id="restrictions_section"></div>', unsafe_allow_html=True)
    st.subheader("Restricciones Temporales Activas")
    restrictions_path = os.path.join(aircraft_folder, "MD_LD_BULK_restrictions.csv")
    if not os.path.exists(restrictions_path):
        st.error(f"No se encontró el archivo en: {restrictions_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        return
    restricciones_df = pd.read_csv(restrictions_path, sep=";", decimal=",")
    restricciones_df.columns = [col.strip().replace(" ", "_") for col in restricciones_df.columns]
    restricciones_df["Temp_Restriction_Symmetric"] = pd.to_numeric(restricciones_df["Temp_Restriction_Symmetric"], errors="coerce").fillna(0)
    restricciones_df["Temp_Restriction_Asymmetric"] = pd.to_numeric(restricciones_df["Temp_Restriction_Asymmetric"], errors="coerce").fillna(0)

    active_restrictions = restricciones_df[
        (restricciones_df["Temp_Restriction_Symmetric"] != 0) | (restricciones_df["Temp_Restriction_Asymmetric"] != 0)
    ][["Position", "Bodega", "Temp_Restriction_Symmetric", "Temp_Restriction_Asymmetric"]]
    
    if active_restrictions.empty:
        st.info(f"No hay restricciones temporales activas para la aeronave {tail}.")
    else:
        st.write(f"Restricciones temporales activas para la aeronave {tail}:")
        st.dataframe(
            active_restrictions,
            column_config={
                "Position": "Posición",
                "Bodega": "Bodega",
                "Temp_Restriction_Symmetric": "Restricción Temporal Simétrica (kg)",
                "Temp_Restriction_Asymmetric": "Restricción Temporal Asimétrica (kg)"
            },
            use_container_width=True
        )

    exclusions_path = os.path.join(aircraft_folder, "exclusiones.csv")
    if not os.path.exists(exclusions_path):
        st.error(f"No se encontró el archivo en: {exclusions_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        return
    exclusiones_df = pd.read_csv(exclusions_path, sep=";", decimal=",")
    exclusiones_df.set_index(exclusiones_df.columns[0], inplace=True)

    cumulative_restrictions_aft_path = os.path.join(aircraft_folder, "cummulative_restrictions_AFT.csv")
    if not os.path.exists(cumulative_restrictions_aft_path):
        st.error(f"No se encontró el archivo en: {cumulative_restrictions_aft_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        return
    cumulative_restrictions_aft_df = pd.read_csv(cumulative_restrictions_aft_path, sep=";", decimal=",")

    cumulative_restrictions_fwd_path = os.path.join(aircraft_folder, "cummulative_restrictions_FWD.csv")
    if not os.path.exists(cumulative_restrictions_fwd_path):
        st.error(f"No se encontró el archivo en: {cumulative_restrictions_fwd_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        return
    cumulative_restrictions_fwd_df = pd.read_csv(cumulative_restrictions_fwd_path, sep=";", decimal=",")

    fuel_table_path = os.path.join(aircraft_folder, "Usable_fuel_table.csv")
    if not os.path.exists(fuel_table_path):
        st.error(f"No se encontró el archivo en: {fuel_table_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        return
    fuel_table = pd.read_csv(fuel_table_path, sep=";", decimal=",", encoding="latin-1")
    required_fuel_columns = ["Fuel_kg", "Outer Tank LH", "Outer Tank RH", "Inner Tank LH", "Inner Tank RH", "Central Tank", "Trim Tank", "MOMENT-X", "MOMENT-Y"]
    if not all(col in fuel_table.columns for col in required_fuel_columns):
        st.error(f"El archivo Usable_fuel_table.csv no contiene las columnas esperadas: {required_fuel_columns}.")
        return

    outer_tanks_path = os.path.normpath(os.path.join(aircraft_folder, "outer_tanks.csv"))
    inner_tanks_path = os.path.normpath(os.path.join(aircraft_folder, "inner_tanks.csv"))
    center_tank_path = os.path.normpath(os.path.join(aircraft_folder, "center_tank.csv"))
    trim_tank_path = os.path.normpath(os.path.join(aircraft_folder, "trim_tank.csv"))

    missing_files = []
    if not os.path.exists(outer_tanks_path):
        missing_files.append("outer_tanks.csv")
    if not os.path.exists(inner_tanks_path):
        missing_files.append("inner_tanks.csv")
    if not os.path.exists(center_tank_path):
        missing_files.append("center_tank.csv")
    if not os.path.exists(trim_tank_path):
        missing_files.append("trim_tank.csv")

    if missing_files:
        st.error(f"Faltan los siguientes archivos en la carpeta de la aeronave: {', '.join(missing_files)}")
        return

    outer_tanks_df = pd.read_csv(outer_tanks_path, sep=";", decimal=",")
    inner_tanks_df = pd.read_csv(inner_tanks_path, sep=";", decimal=",")
    center_tank_df = pd.read_csv(center_tank_path, sep=";", decimal=",")
    trim_tank_df = pd.read_csv(trim_tank_path, sep=";", decimal=",")

    passengers_path = os.path.join(aircraft_folder, "Passengers.csv")
    if not os.path.exists(passengers_path):
        st.error(f"No se encontró el archivo en: {passengers_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        return
    passengers_df = pd.read_csv(passengers_path, sep=";", decimal=",")
    max_passengers_supernumerary = int(passengers_df["Quantity-Passenger"].max())
    if 0 not in passengers_df["Quantity-Passenger"].values:
        passengers_df = pd.concat([pd.DataFrame({"Quantity-Passenger": [0], "Weight": [0], "Moment": [0]}), passengers_df], ignore_index=True)

    flite_deck_path = os.path.join(aircraft_folder, "Flite_deck_passengers.csv")
    if not os.path.exists(flite_deck_path):
        st.error(f"No se encontró el archivo en: {flite_deck_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        return
    flite_deck_df = pd.read_csv(flite_deck_path, sep=";", decimal=",")
    max_passengers_cockpit = int(flite_deck_df["Quantity-Passenger Flite-Deck"].max())
    if 0 not in flite_deck_df["Quantity-Passenger Flite-Deck"].values:
        flite_deck_df = pd.concat([pd.DataFrame({"Quantity-Passenger Flite-Deck": [0], "Weight": [0], "Moment": [0]}), flite_deck_df], ignore_index=True)

    trimset_path = os.path.join(aircraft_folder, "trimset.csv")
    if not os.path.exists(trimset_path):
        st.error(f"No se encontró el archivo en: {trimset_path}. Asegúrate de que el archivo exista en la ruta especificada.")
        return
    trimset_df = pd.read_csv(trimset_path, sep=";", decimal=",")

    st.markdown('<div id="flight_info_section"></div>', unsafe_allow_html=True)
    st.subheader("Información del Vuelo")
    
    col_fuel1, col_fuel2 = st.columns(2)
    with col_fuel1:
        normal_fuel = st.number_input("Combustible Total (kg)", min_value=0.0, value=st.session_state.get("normal_fuel", float(default_flight_data["fuel_kg"])), key="normal_fuel")
    with col_fuel2:
        computed_ballast_fuel = st.session_state.get("computed_ballast_fuel", float(default_flight_data["ballast_fuel"]))
        st.number_input(
            "Combustible Ballast y/o Atrapado (kg)",
            min_value=0.0,
            value=computed_ballast_fuel,
            key="ballast_fuel",
            disabled=True
        )

    # Definir fuel_kg para el modo Automático
    fuel_kg = normal_fuel + computed_ballast_fuel

    col1, col2, col3 = st.columns(3)
    with col1:
        fuel_mode = st.selectbox("Método de Cargue de Combustible", ["Automático", "Manual"], index=["Automático", "Manual"].index(st.session_state.get("fuel_mode", default_calc_state["fuel_mode"])), key="fuel_mode")
    with col2:
        trip_fuel = st.number_input("Trip Fuel (kg)", min_value=0.0, value=st.session_state.get("trip_fuel", float(default_flight_data["trip_fuel"])), key="trip_fuel")
    with col3:
        taxi_fuel = st.number_input("Taxi Fuel (kg)", min_value=0.0, value=st.session_state.get("taxi_fuel", float(default_flight_data["taxi_fuel"])), key="taxi_fuel")

    tank_fuel = st.session_state.get("fuel_distribution", default_calc_state["fuel_distribution"]).copy()
    if fuel_mode == "Manual":
        st.write("### Cargue Manual de Combustible")
        st.write("Ingrese la cantidad de combustible (kg) en cada tanque. El Combustible Total excluye el combustible atrapado si 'Atrapado' está marcado.")

        tanks = {
            "Outer Tank LH": {"df": outer_tanks_df, "max_kg": 2850},
            "Outer Tank RH": {"df": outer_tanks_df, "max_kg": 2850},
            "Inner Tank LH": {"df": inner_tanks_df, "max_kg": 32950},
            "Inner Tank RH": {"df": inner_tanks_df, "max_kg": 32950},
            "Center Tank": {"df": center_tank_df, "max_kg": 32725},
            "Trim Tank": {"df": trim_tank_df, "max_kg": 4875}
        }

        if tail != "N342AV":
            tanks["Center Tank"]["max_kg"] = 0
            tank_fuel["Center Tank"] = 0.0
            st.warning("El tanque central no está disponible para esta aeronave.")

        tank_inputs = st.columns(3)
        for idx, tank in enumerate(tanks.keys()):
            with tank_inputs[idx % 3]:
                if tank == "Center Tank" and tail != "N342AV":
                    st.number_input(
                        f"{tank} (kg, máx {tanks[tank]['max_kg']})",
                        min_value=0.0,
                        max_value=0.0,
                        value=0.0,
                        disabled=True,
                        key=f"tank_{tank}"
                    )
                else:
                    if tank == "Trim Tank":
                        col_tank, col_checkbox = st.columns([3, 1])
                        with col_tank:
                            tank_fuel[tank] = st.number_input(
                                f"{tank} (kg, máx {tanks[tank]['max_kg']})",
                                min_value=0.0,
                                max_value=float(tanks[tank]['max_kg']),
                                value=st.session_state.get(f"tank_{tank}", float(tank_fuel[tank])),
                                key=f"tank_{tank}"
                            )
                        with col_checkbox:
                            trapped = st.checkbox("Atrapado", value=st.session_state.get("trapped_fuel", False), key="trapped_fuel")
                            st.session_state.computed_ballast_fuel = tank_fuel[tank] if trapped else 0.0
                    else:
                        tank_fuel[tank] = st.number_input(
                            f"{tank} (kg, máx {tanks[tank]['max_kg']})",
                            min_value=0.0,
                            max_value=float(tanks[tank]['max_kg']),
                            value=st.session_state.get(f"tank_{tank}", float(tank_fuel[tank])),
                            key=f"tank_{tank}"
                        )

        total_fuel_input = sum(tank_fuel.values())
        fuel_kg = total_fuel_input - st.session_state.computed_ballast_fuel
        
        st.info(f"Combustible Total (excluyendo atrapado): {fuel_kg:.1f} kg")
        if st.session_state.computed_ballast_fuel > 0:
            st.info(f"Combustible Atrapado: {st.session_state.computed_ballast_fuel:.1f} kg")
        else:
            st.info("No hay combustible atrapado.")

        non_trapped_fuel = sum(tank_fuel[tank] for tank in tanks if tank != "Trim Tank" or not trapped)
        if abs(non_trapped_fuel - fuel_kg) > 0.01:
            st.error(f"Error interno: La suma de los tanques no atrapados ({non_trapped_fuel:.1f} kg) no coincide con fuel_kg ({fuel_kg:.1f} kg).")
        else:
            st.success(f"La suma de los tanques no atrapados ({non_trapped_fuel:.1f} kg) coincide con el Combustible Total ({fuel_kg:.1f} kg).")

        if trip_fuel > fuel_kg:
            st.error("El Trip Fuel no puede ser mayor que el Combustible Total (excluyendo atrapado).")
        if taxi_fuel > (fuel_kg - trip_fuel):
            st.error("El Taxi Fuel no puede ser mayor que el combustible disponible después del Trip Fuel.")
    
    col4, col5, col6 = st.columns(3)
    with col4:
        tipo_carga = st.selectbox("Tipo de cargue", ["Simétrico", "Asimétrico"], index=["Simétrico", "Asimétrico"].index(st.session_state.get("tipo_carga", default_flight_data["tipo_carga"])), key="tipo_carga")
    with col5:
        destino_inicial = st.text_input("Destino inicial (ej. MIA)", value=st.session_state.get("destino_inicial", default_flight_data["destino_inicial"]), key="destino_inicial").upper()
    with col6:
        takeoff_runway = st.text_input("Pista de despegue (ej. RWY 13)", value=st.session_state.get("takeoff_runway", default_flight_data["takeoff_runway"]), key="takeoff_runway")

    col7, col8, col9 = st.columns(3)
    with col7:
        rwy_condition = st.selectbox("Condición de la pista", ["Dry", "Wet", "Contaminated"], index=["Dry", "Wet", "Contaminated"].index(st.session_state.get("rwy_condition", default_flight_data["rwy_condition"])), key="rwy_condition")
    with col8:
        flaps_conf = st.selectbox("Configuración de flaps", ["1+F", "2", "3"], index=["1+F", "2", "3"].index(st.session_state.get("flaps_conf", default_flight_data["flaps_conf"])), key="flaps_conf")
    with col9:
        temperature = st.number_input("Temperatura (°C)", value=st.session_state.get("temperature", float(default_flight_data["temperature"])), key="temperature")

    col10, col11, col12 = st.columns(3)
    with col10:
        air_condition = st.selectbox("Packs", ["On", "Off"], index=["On", "Off"].index(st.session_state.get("air_condition", default_flight_data["air_condition"])), key="air_condition")
    with col11:
        anti_ice = st.selectbox("Anti ice", ["On", "Off"], index=["On", "Off"].index(st.session_state.get("anti_ice", default_flight_data["anti_ice"])), key="anti_ice")
    with col12:
        qnh = st.number_input("QNH (hPa)", min_value=900.0, max_value=1100.0, value=st.session_state.get("qnh", float(default_flight_data["qnh"])), key="qnh")

    col13, col14, col15 = st.columns(3)
    with col13:
        performance_tow = st.number_input("Performance TOW (kg)", min_value=0.0, value=st.session_state.get("performance_tow", float(default_flight_data["performance_tow"])), key="performance_tow")
    with col14:
        performance_lw = st.number_input("Performance LW (kg)", min_value=0.0, value=st.session_state.get("performance_lw", float(default_flight_data["performance_lw"])), key="performance_lw")
    with col15:
        passengers_cockpit = st.number_input(f"Pasajeros en cabina de mando (máx {max_passengers_cockpit})", min_value=0, max_value=max_passengers_cockpit, step=1, value=st.session_state.get("passengers_cockpit", int(default_flight_data["passengers_cockpit"])), key="passengers_cockpit")

    col16, col17, _ = st.columns(3)
    with col16:
        passengers_supernumerary = st.number_input(f"Pasajeros supernumerarios (máx {max_passengers_supernumerary})", min_value=0, max_value=max_passengers_supernumerary, step=1, value=st.session_state.get("passengers_supernumerary", int(default_flight_data["passengers_supernumerary"])), key="passengers_supernumerary")

    if fuel_kg < 0 or taxi_fuel < 0 or trip_fuel < 0:
        st.error("Los valores de combustible no pueden ser negativos.")
        return
    if trip_fuel > (fuel_kg - taxi_fuel):
        st.error("El Trip Fuel no puede ser mayor que el combustible disponible después del Taxi Fuel.")
        return

    aircraft_data = AircraftData(
        tail=tail,
        mtoc=basic_data["MTOW (kg)"].values[0],
        mlw=basic_data["MLW"].values[0],
        mzfw=basic_data["MZFW"].values[0],
        oew=basic_data["OEW"].values[0],
        arm=basic_data["ARM"].values[0],
        moment_aircraft=basic_data["Moment_Aircraft"].values[0],
        cg_aircraft=basic_data["CG_Aircraft"].values[0],
        lemac=basic_data["LEMAC"].values[0],
        mac_length=basic_data["MAC_length"].values[0],
        mrw_limit=basic_data["MRW"].values[0],
        lateral_imbalance_limit=basic_data["Lateral_Imbalance_Limit"].values[0],
        ldf_limit=basic_data["LDF_LIMIT"].values[0],
        lda_limit=basic_data["LDA_LIMIT"].values[0]
    )

    if aircraft_data.mac_length == 0:
        st.error("MAC_length no puede ser cero.")
        return

    if aircraft_data.lemac == 0:
        st.error("LEMAC no puede ser cero.")
        return

    passenger_cockpit_row = flite_deck_df[flite_deck_df["Quantity-Passenger Flite-Deck"] == passengers_cockpit].iloc[0]
    passengers_cockpit_total_weight = passenger_cockpit_row["Weight"]
    passengers_cockpit_total_moment_x = passenger_cockpit_row["Moment"]

    passenger_supernumerary_row = passengers_df[passengers_df["Quantity-Passenger"] == passengers_supernumerary].iloc[0]
    passengers_supernumerary_total_weight = passenger_supernumerary_row["Weight"]
    passengers_supernumerary_total_moment_x = passenger_supernumerary_row["Moment"]

    bow = aircraft_data.oew + passengers_cockpit_total_weight + passengers_supernumerary_total_weight
    bow_moment_x = aircraft_data.moment_aircraft + passengers_cockpit_total_moment_x + passengers_supernumerary_total_moment_x
    bow_moment_y = 0

    fuel_for_tow = fuel_kg - taxi_fuel
    fuel_for_lw = fuel_kg - taxi_fuel - trip_fuel

    if fuel_mode == "Automático":
        required_columns = [
            "Fuel_kg", "Outer Tank LH", "Outer Tank RH", "Inner Tank LH",
            "Inner Tank RH", "Central Tank", "Trim Tank", "MOMENT-X", "MOMENT-Y"
        ]
        missing_columns = [col for col in required_columns if col not in fuel_table.columns]
        if missing_columns:
            st.error(f"Faltan columnas en Usable_fuel_table.csv: {', '.join(missing_columns)}")
            return
        
        fuel_row_tow = fuel_table.iloc[(fuel_table["Fuel_kg"] - fuel_for_tow).abs().argsort()[0]]
        moment_x_fuel_tow = fuel_row_tow["MOMENT-X"]
        moment_y_fuel_tow = fuel_row_tow["MOMENT-Y"]
        tank_fuel = {
            "Outer Tank LH": fuel_row_tow["Outer Tank LH"],
            "Outer Tank RH": fuel_row_tow["Outer Tank RH"],
            "Inner Tank LH": fuel_row_tow["Inner Tank LH"],
            "Inner Tank RH": fuel_row_tow["Inner Tank RH"],
            "Center Tank": fuel_row_tow["Central Tank"],
            "Trim Tank": fuel_row_tow["Trim Tank"]
        }
    else:
        moment_x_fuel_tow = 0.0
        moment_y_fuel_tow = 0.0
        tanks = {
            "Outer Tank LH": {"df": outer_tanks_df, "max_kg": 2850},
            "Outer Tank RH": {"df": outer_tanks_df, "max_kg": 2850},
            "Inner Tank LH": {"df": inner_tanks_df, "max_kg": 32950},
            "Inner Tank RH": {"df": inner_tanks_df, "max_kg": 32950},
            "Center Tank": {"df": center_tank_df, "max_kg": 32725},
            "Trim Tank": {"df": trim_tank_df, "max_kg": 4875}
        }
        for tank, fuel in tank_fuel.items():
            if fuel > 0:
                tank_df = tanks[tank]["df"]
                closest_row = tank_df.iloc[(tank_df["Kg_Fuel"] - fuel).abs().argsort()[0]]
                closest_fuel = closest_row["Kg_Fuel"]
                ratio = fuel / closest_fuel if closest_fuel != 0 else 0
                if tank == "Outer Tank LH":
                    moment_x_fuel_tow += closest_row["Moment_X_OLH"] * ratio
                    moment_y_fuel_tow += closest_row["Moment_Y_OLH"] * ratio
                elif tank == "Outer Tank RH":
                    moment_x_fuel_tow += closest_row["Moment_X_ORH"] * ratio
                    moment_y_fuel_tow += closest_row["Moment_Y_ORH"] * ratio
                elif tank == "Inner Tank LH":
                    moment_x_fuel_tow += closest_row["Moment_X_ILH"] * ratio
                    moment_y_fuel_tow += closest_row["Moment_Y_ILH"] * ratio
                elif tank == "Inner Tank RH":
                    moment_x_fuel_tow += closest_row["Moment_X_IRH"] * ratio
                    moment_y_fuel_tow += closest_row["Moment_Y_IRH"] * ratio
                elif tank == "Center Tank":
                    moment_x_fuel_tow += closest_row["CT_MOMENT_X"] * ratio
                elif tank == "Trim Tank":
                    moment_x_fuel_tow += closest_row["T_MOMENT_X"] * ratio

    if fuel_for_lw > 0:
        fuel_per_inner_tank = fuel_for_lw / 2
        max_inner_tank_capacity = 32950

        if fuel_per_inner_tank > max_inner_tank_capacity:
            st.warning(f"El combustible por tanque interno ({fuel_per_inner_tank:.1f} kg) excede la capacidad máxima ({max_inner_tank_capacity:.1f} kg). Se usará el valor máximo.")
            fuel_per_inner_tank = max_inner_tank_capacity

        inner_tank_row = inner_tanks_df.iloc[(inner_tanks_df["Kg_Fuel"] - fuel_per_inner_tank).abs().argsort()[0]]
        closest_fuel = inner_tank_row["Kg_Fuel"]
        ratio = fuel_per_inner_tank / closest_fuel if closest_fuel != 0 else 0

        moment_x_fuel_lw = (inner_tank_row["Moment_X_ILH"] + inner_tank_row["Moment_X_IRH"]) * ratio
        moment_y_fuel_lw = (inner_tank_row["Moment_Y_ILH"] + inner_tank_row["Moment_Y_IRH"]) * ratio
    else:
        moment_x_fuel_lw = 0.0
        moment_y_fuel_lw = 0.0

    st.markdown('<div id="manifest_section"></div>', unsafe_allow_html=True)
    st.subheader("Carga del Manifiesto")
    manifiesto_option = st.radio("Seleccione cómo ingresar el manifiesto", ["Ingresar Manualmente", "Subir CSV"], index=0)

    if "calculation_state" not in st.session_state:
        st.session_state.calculation_state = CalculationState(**default_calc_state)

    operador = st.session_state.get("operador_manual", default_flight_data["operador"])
    numero_vuelo = st.session_state.get("numero_vuelo_manual", default_flight_data["numero_vuelo"])
    matricula = st.session_state.get("matricula_manual", default_flight_data["matricula"] or tail)
    fecha_vuelo = st.session_state.get("fecha_vuelo_manual", default_flight_data["fecha_vuelo"] or datetime.now().strftime("%d/%m/%Y"))
    hora_vuelo = st.session_state.get("hora_vuelo_manual", default_flight_data["hora_vuelo"] or datetime.now().strftime("%H:%M"))
    ruta_vuelo = st.session_state.get("ruta_vuelo_manual", default_flight_data["ruta_vuelo"] or "Ruta Desconocida")
    revision = st.session_state.get("revision_manual", default_flight_data["revision"] or "0")
    fecha_vuelo_safe = fecha_vuelo.replace("/", "_")

    if manifiesto_option == "Subir CSV":
        st.markdown('<div id="manifest_data_section"></div>', unsafe_allow_html=True)
        manifiesto_file = st.file_uploader("Sube el manifiesto CSV", type="csv", key="manifiesto")
        if manifiesto_file:
            df = pd.read_csv(manifiesto_file, skiprows=8, sep=";", encoding="latin-1", header=None, decimal=",")
            df.columns = ["Contour", "Number ULD", "ULD Final Destination", "Weight (KGS)", "Pieces", "Notes", "Extra1", "Extra2", "Extra3", "Extra4"]
            df = df[["Contour", "Number ULD", "ULD Final Destination", "Weight (KGS)", "Pieces", "Notes"]]
            df = df.dropna(subset=["Number ULD", "Weight (KGS)"], how="any")
            df = df[~df["Number ULD"].astype(str).str.upper().str.contains("TOTAL")]
            df = df[~df["Contour"].astype(str).str.upper().str.contains("TOTAL")]
            df["Weight (KGS)"] = pd.to_numeric(df["Weight (KGS)"], errors="coerce")
            
            df[["Pallet Base Size", "Baseplate Code"]] = df["Number ULD"].apply(lambda x: pd.Series(clasificar_base_refinada(x)))
            df["Posiciones Sugeridas"] = df.apply(lambda row: sugerencias_final_con_fak(row, restricciones_df, tipo_carga.lower()), axis=1)
            df["Posición Asignada"] = ""
            df["X-arm"] = None
            df["Y-arm"] = None
            df["Momento X"] = None
            df["Momento Y"] = None
            df["Bodega"] = None
            df["Rotated"] = False
            
            if st.session_state.calculation_state.df is not None:
                prev_df = st.session_state.calculation_state.df
                if prev_df[["Number ULD", "Weight (KGS)"]].equals(df[["Number ULD", "Weight (KGS)"]]):
                    df.update(prev_df[["Posición Asignada", "X-arm", "Y-arm", "Momento X", "Momento Y", "Bodega", "Rotated"]])
                    st.session_state.calculation_state.posiciones_usadas = set(df[df["Posición Asignada"] != ""]["Posición Asignada"].tolist())
                    st.session_state.calculation_state.rotaciones = {row["Number ULD"]: row["Rotated"] for _, row in df[df["Rotated"] != False].iterrows()}
                else:
                    st.session_state.calculation_state.posiciones_usadas = set()
                    st.session_state.calculation_state.rotaciones = {}
            
            st.session_state.calculation_state.df = df.copy()
            st.session_state.manifiesto_manual = df.copy()
            st.write("Manifiesto Inicial:")
            st.dataframe(
                df,
                column_config={
                    "Contour": st.column_config.TextColumn("Contour"),
                    "Number ULD": st.column_config.TextColumn("Number ULD"),
                    "ULD Final Destination": st.column_config.TextColumn("ULD Final Destination"),
                    "Weight (KGS)": st.column_config.NumberColumn("Weight (KGS)", format="%.1f"),
                    "Pieces": st.column_config.NumberColumn("Pieces"),
                    "Notes": st.column_config.TextColumn("Notes"),
                    "Pallet Base Size": st.column_config.TextColumn("Pallet Base Size"),
                    "Baseplate Code": st.column_config.TextColumn("Baseplate Code"),
                    "Posiciones Sugeridas": st.column_config.ListColumn("Posiciones Sugeridas"),
                    "Posición Asignada": st.column_config.TextColumn("Posición Asignada"),
                    "X-arm": st.column_config.NumberColumn("X-arm", format="%.3f"),
                    "Y-arm": st.column_config.NumberColumn("Y-arm", format="%.3f"),
                    "Momento X": st.column_config.NumberColumn("Momento X", format="%.3f"),
                    "Momento Y": st.column_config.NumberColumn("Momento Y", format="%.3f"),
                    "Bodega": st.column_config.TextColumn("Bodega"),
                    "Rotated": st.column_config.CheckboxColumn("Rotated")
                },
                use_container_width=True
            )

            manifiesto_file.seek(0)
            content = manifiesto_file.read().decode("latin-1")
            lines = content.splitlines()[:10]
            lines = [line.strip().split(";") for line in lines]
            
            operador = lines[2][0]
            revision = lines[4][6]
            fecha_vuelo = lines[6][1]
            hora_vuelo = lines[6][3]
            ruta_vuelo = lines[6][5]
            matricula = lines[6][7]
            numero_vuelo = lines[7][1]
            fecha_vuelo_safe = fecha_vuelo.replace("/", "_")
        elif st.session_state.calculation_state.df is not None:
            st.write("Manifiesto cargado previamente:")
            st.dataframe(
                st.session_state.calculation_state.df,
                column_config={
                    "Contour": st.column_config.TextColumn("Contour"),
                    "Number ULD": st.column_config.TextColumn("Number ULD"),
                    "ULD Final Destination": st.column_config.TextColumn("ULD Final Destination"),
                    "Weight (KGS)": st.column_config.NumberColumn("Weight (KGS)", format="%.1f"),
                    "Pieces": st.column_config.NumberColumn("Pieces"),
                    "Notes": st.column_config.TextColumn("Notes"),
                    "Pallet Base Size": st.column_config.TextColumn("Pallet Base Size"),
                    "Baseplate Code": st.column_config.TextColumn("Baseplate Code"),
                    "Posiciones Sugeridas": st.column_config.ListColumn("Posiciones Sugeridas"),
                    "Posición Asignada": st.column_config.TextColumn("Posición Asignada"),
                    "X-arm": st.column_config.NumberColumn("X-arm", format="%.3f"),
                    "Y-arm": st.column_config.NumberColumn("Y-arm", format="%.3f"),
                    "Momento X": st.column_config.NumberColumn("Momento X", format="%.3f"),
                    "Momento Y": st.column_config.NumberColumn("Momento Y", format="%.3f"),
                    "Bodega": st.column_config.TextColumn("Bodega"),
                    "Rotated": st.column_config.CheckboxColumn("Rotated")
                },
                use_container_width=True
            )
    else:
        st.markdown('<div id="manifest_flight_info_section"></div>', unsafe_allow_html=True)
        st.subheader("Información del Vuelo para el Manifiesto")
        st.write("Ingrese los detalles del vuelo para el manifiesto manual.")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            operador_manual = st.text_input("Operador", value=st.session_state.get("operador_manual", default_flight_data["operador"]), key="operador_manual")
            numero_vuelo_manual = st.text_input("Número de Vuelo", value=st.session_state.get("numero_vuelo_manual", default_flight_data["numero_vuelo"]), key="numero_vuelo_manual")
        with col2:
            matricula_manual = st.text_input("Matrícula", value=st.session_state.get("matricula_manual", default_flight_data["matricula"] or tail), key="matricula_manual")
            fecha_vuelo_manual = st.text_input("Fecha (DD/MM/YYYY)", value=st.session_state.get("fecha_vuelo_manual", default_flight_data["fecha_vuelo"] or datetime.now().strftime("%d/%m/%Y")), key="fecha_vuelo_manual")
        with col3:
            hora_vuelo_manual = st.text_input("Hora (HH:MM)", value=st.session_state.get("hora_vuelo_manual", default_flight_data["hora_vuelo"] or datetime.now().strftime("%H:%M")), key="hora_vuelo_manual")
            ruta_vuelo_manual = st.text_input("Ruta", value=st.session_state.get("ruta_vuelo_manual", default_flight_data["ruta_vuelo"] or "ORI-DES"), key="ruta_vuelo_manual")
        
        revision_manual = st.text_input("Revisión de Manifiesto", value=st.session_state.get("revision_manual", default_flight_data["revision"] or "0"), key="revision_manual")
        
        st.markdown('<div id="manifest_data_section"></div>', unsafe_allow_html=True)
        st.subheader("Datos del Manifiesto")
        st.write("Ingrese los datos del manifiesto en la tabla siguiente:")
        edited_df = st.data_editor(
            st.session_state.manifiesto_manual,
            column_config={
                "Contour": st.column_config.TextColumn("Contour"),
                "Number ULD": st.column_config.TextColumn("Number ULD", required=True),
                "ULD Final Destination": st.column_config.TextColumn("ULD Final Destination"),
                "Weight (KGS)": st.column_config.NumberColumn("Weight (KGS)", min_value=0.0, required=True),
                "Pieces": st.column_config.NumberColumn("Pieces", min_value=0, step=1),
                "Notes": st.column_config.TextColumn("Notes")
            },
            num_rows="dynamic",
            use_container_width=True,
            key=f"manifiesto_editor_{st.session_state.get('edit_count', 0)}"
        )
        
        if st.button("Confirmar Manifiesto Manual", key="confirm_manifest"):
            if edited_df.empty or edited_df[["Number ULD", "Weight (KGS)"]].isna().any().any():
                st.error("El manifiesto no puede estar vacío y debe incluir 'Number ULD' y 'Weight (KGS)' para cada fila.")
            else:
                df = edited_df.copy()
                df = df.dropna(subset=["Number ULD", "Weight (KGS)"], how="any")
                df = df[~(df["Number ULD"].astype(str).str.upper().str.contains("TOTAL") | df["Contour"].astype(str).str.upper().str.contains("TOTAL"))]
                df["Weight (KGS)"] = pd.to_numeric(df["Weight (KGS)"], errors="coerce")
                
                df[["Pallet Base Size", "Baseplate Code"]] = df["Number ULD"].apply(lambda x: pd.Series(clasificar_base_refinada(x)))
                df["Posiciones Sugeridas"] = df.apply(lambda row: sugerencias_final_con_fak(row, restricciones_df, tipo_carga.lower()), axis=1)
                df["Posición Asignada"] = ""
                df["X-arm"] = None
                df["Y-arm"] = None
                df["Momento X"] = None
                df["Momento Y"] = None
                df["Bodega"] = None
                df["Rotated"] = False
                
                if st.session_state.calculation_state.df is not None:
                    prev_df = st.session_state.calculation_state.df
                    df["key"] = df["Number ULD"].astype(str) + "_" + df["Weight (KGS)"].astype(str)
                    prev_df["key"] = prev_df["Number ULD"].astype(str) + "_" + prev_df["Weight (KGS)"].astype(str)
                    for idx, row in df.iterrows():
                        matching_rows = prev_df[prev_df["key"] == row["key"]]
                        if not matching_rows.empty:
                            df.loc[idx, ["Posición Asignada", "X-arm", "Y-arm", "Momento X", "Momento Y", "Bodega", "Rotated"]] = matching_rows.iloc[0][["Posición Asignada", "X-arm", "Y-arm", "Momento X", "Momento Y", "Bodega", "Rotated"]]
                    df.drop(columns=["key"], inplace=True)
                    if "key" in prev_df.columns:
                        prev_df.drop(columns=["key"], inplace=True)
                
                st.session_state.calculation_state.df = df.copy()
                st.session_state.manifiesto_manual = df.copy()
                st.write("Manifiesto Ingresado:")
                st.dataframe(
                    df,
                    column_config={
                        "Contour": st.column_config.TextColumn("Contour"),
                        "Number ULD": st.column_config.TextColumn("Number ULD"),
                        "ULD Final Destination": st.column_config.TextColumn("ULD Final Destination"),
                        "Weight (KGS)": st.column_config.NumberColumn("Weight (KGS)", format="%.1f"),
                        "Pieces": st.column_config.NumberColumn("Pieces"),
                        "Notes": st.column_config.TextColumn("Notes"),
                        "Pallet Base Size": st.column_config.TextColumn("Pallet Base Size"),
                        "Baseplate Code": st.column_config.TextColumn("Baseplate Code"),
                        "Posiciones Sugeridas": st.column_config.ListColumn("Posiciones Sugeridas"),
                        "Posición Asignada": st.column_config.TextColumn("Posición Asignada"),
                        "X-arm": st.column_config.NumberColumn("X-arm", format="%.3f"),
                        "Y-arm": st.column_config.NumberColumn("Y-arm", format="%.3f"),
                        "Momento X": st.column_config.NumberColumn("Momento X", format="%.3f"),
                        "Momento Y": st.column_config.NumberColumn("Momento Y", format="%.3f"),
                        "Bodega": st.column_config.TextColumn("Bodega"),
                        "Rotated": st.column_config.CheckboxColumn("Rotated")
                    },
                    use_container_width=True
                )
                
                operador = operador_manual
                numero_vuelo = numero_vuelo_manual
                matricula = matricula_manual
                fecha_vuelo = fecha_vuelo_manual
                hora_vuelo = hora_vuelo_manual
                ruta_vuelo = ruta_vuelo_manual
                revision = revision_manual
                fecha_vuelo_safe = fecha_vuelo.replace("/", "_")

    df = st.session_state.calculation_state.df

    if df is None:
        st.warning("Por favor, suba un manifiesto CSV o confirme un manifiesto manual para continuar.")
        return

    flight_data = FlightData(
        operador=operador,
        numero_vuelo=numero_vuelo,
        matricula=matricula,
        fecha_vuelo=fecha_vuelo,
        hora_vuelo=hora_vuelo,
        ruta_vuelo=ruta_vuelo,
        revision=revision,
        destino_inicial=destino_inicial,
        fuel_kg=fuel_kg,
        trip_fuel=trip_fuel,
        taxi_fuel=taxi_fuel,
        tipo_carga=tipo_carga.lower(),
        takeoff_runway=takeoff_runway,
        rwy_condition=rwy_condition,
        flaps_conf=flaps_conf,
        temperature=temperature,
        air_condition=air_condition,
        anti_ice=anti_ice,
        qnh=qnh,
        performance_tow=performance_tow,
        performance_lw=performance_lw,
        passengers_cockpit=passengers_cockpit,
        passengers_supernumerary=passengers_supernumerary
    )

    st.session_state.calculation_state.bow = bow
    st.session_state.calculation_state.bow_moment_x = bow_moment_x
    st.session_state.calculation_state.bow_moment_y = bow_moment_y
    st.session_state.calculation_state.moment_x_fuel_tow = moment_x_fuel_tow
    st.session_state.calculation_state.moment_y_fuel_tow = moment_y_fuel_tow
    st.session_state.calculation_state.moment_x_fuel_lw = moment_x_fuel_lw
    st.session_state.calculation_state.moment_y_fuel_lw = moment_y_fuel_lw
    st.session_state.calculation_state.passengers_cockpit_total_weight = passengers_cockpit_total_weight
    st.session_state.calculation_state.passengers_cockpit_total_moment_x = passengers_cockpit_total_moment_x
    st.session_state.calculation_state.passengers_supernumerary_total_weight = passengers_supernumerary_total_weight
    st.session_state.calculation_state.passengers_supernumerary_total_moment_x = passengers_supernumerary_total_moment_x
    st.session_state.calculation_state.fuel_distribution = tank_fuel
    st.session_state.calculation_state.fuel_mode = fuel_mode

    st.markdown('<div id="weight_restrictions_debug_section"></div>', unsafe_allow_html=True)
    st.subheader("Restricciones de Peso por Posición")
    with st.expander("Ver Restricciones de Peso por Posición", expanded=False):
        st.write(f"Tipo de carga seleccionado: {tipo_carga}")

        if restricciones_df is None or restricciones_df.empty:
            st.error("Error: No se pudo cargar MD_LD_BULK_restrictions.csv")
        else:
            numeric_cols = [
                "Symmetric_Max_Weight_(kg)_5%", "Asymmetric_Max_Weight_(kg)_5%",
                "Temp_Restriction_Symmetric", "Temp_Restriction_Asymmetric"
            ]
            for col in numeric_cols:
                restricciones_df[col] = pd.to_numeric(restricciones_df[col], errors="coerce").fillna(0)
                if restricciones_df[col].isna().any():
                    st.warning(f"Advertencia: Valores nulos en {col} en MD_LD_BULK_restrictions.csv")

            debug_restrictions = restricciones_df[[
                "Position", "Bodega", "Pallet_Base_size_Allowed",
                "Symmetric_Max_Weight_(kg)_5%", "Temp_Restriction_Symmetric",
                "Asymmetric_Max_Weight_(kg)_5%", "Temp_Restriction_Asymmetric"
            ]].copy()

            debug_restrictions["Peso Máximo Efectivo (kg)"] = debug_restrictions.apply(
                lambda row: calculate_peso_maximo_efectivo(row, tipo_carga),
                axis=1
            )

            debug_restrictions = debug_restrictions.round(2)
            debug_restrictions = debug_restrictions.sort_values(["Bodega", "Position"])

            st.dataframe(
                debug_restrictions,
                use_container_width=True,
                height=400,
                column_config={
                    "Position": st.column_config.TextColumn("Posición", width="medium"),
                    "Bodega": st.column_config.TextColumn("Bodega", width="small"),
                    "Pallet_Base_size_Allowed": st.column_config.TextColumn("Base Permitida", width="medium"),
                    "Symmetric_Max_Weight_(kg)_5%": st.column_config.NumberColumn(
                        "Peso Máx. Simétrico (kg)", format="%.2f", width="medium"
                    ),
                    "Temp_Restriction_Symmetric": st.column_config.NumberColumn(
                        "Restricción Temporal Simétrica (kg)", format="%.2f", width="medium"
                    ),
                    "Asymmetric_Max_Weight_(kg)_5%": st.column_config.NumberColumn(
                        "Peso Máx. Asimétrico (kg)", format="%.2f", width="medium"
                    ),
                    "Temp_Restriction_Asymmetric": st.column_config.NumberColumn(
                        "Restricción Temporal Asimétrica (kg)", format="%.2f", width="medium"
                    ),
                    "Peso Máximo Efectivo (kg)": st.column_config.NumberColumn(
                        "Peso Máximo Efectivo (kg)", format="%.2f", width="medium"
                    ),
                }
            )

            st.write(f"Se cargaron {len(debug_restrictions)} posiciones con restricciones de peso.")
            if debug_restrictions["Peso Máximo Efectivo (kg)"].isna().any():
                st.warning("Advertencia: Algunas posiciones tienen pesos máximos efectivos nulos.")

    st.markdown('<div id="calculation_mode_section"></div>', unsafe_allow_html=True)
    st.subheader("Seleccione el Modo de Cálculo")
    tab1, tab2 = st.tabs(["Cálculo Manual", "Cálculo Automático"])

    with tab1:
        st.markdown('<div id="manual_assignment_section"></div>', unsafe_allow_html=True)
        st.subheader("Asignación Manual de Posiciones")
        manual_assignment(
            st.session_state.calculation_state.df,
            restricciones_df,
            flight_data.tipo_carga,
            exclusiones_df,
            st.session_state.calculation_state.posiciones_usadas,
            st.session_state.calculation_state.rotaciones,
            tab_prefix="manual"
        )

    with tab2:
        automatic_assignment(
            st.session_state.calculation_state.df,
            restricciones_df,
            flight_data.tipo_carga,
            exclusiones_df,
            st.session_state.calculation_state.posiciones_usadas,
            st.session_state.calculation_state.rotaciones,
            flight_data.destino_inicial,
            st.session_state.calculation_state.bow,
            st.session_state.calculation_state.bow_moment_x,
            st.session_state.calculation_state.bow_moment_y,
            flight_data.fuel_kg,
            flight_data.taxi_fuel,
            st.session_state.calculation_state.moment_x_fuel_tow,
            st.session_state.calculation_state.moment_y_fuel_tow,
            aircraft_data.lemac,
            aircraft_data.mac_length,
            cumulative_restrictions_fwd_df,
            cumulative_restrictions_aft_df,
            tab_prefix="auto"
        )

    st.markdown('<div id="desassign_pallets_section"></div>', unsafe_allow_html=True)

    data_to_save = None
    output_json = None

    if st.session_state.calculation_state.df is not None and st.session_state.calculation_state.df["Posición Asignada"].ne("").any():
        df_asignados = st.session_state.calculation_state.df[st.session_state.calculation_state.df["Posición Asignada"] != ""]
        
        final_results = calculate_final_values(
            df_asignados,
            st.session_state.calculation_state.bow,
            st.session_state.calculation_state.bow_moment_x,
            st.session_state.calculation_state.bow_moment_y,
            flight_data.fuel_kg,
            flight_data.taxi_fuel,
            flight_data.trip_fuel,
            st.session_state.calculation_state.moment_x_fuel_tow,
            st.session_state.calculation_state.moment_y_fuel_tow,
            st.session_state.calculation_state.moment_x_fuel_lw,
            st.session_state.calculation_state.moment_y_fuel_lw,
            aircraft_data.lemac,
            aircraft_data.mac_length,
            aircraft_data.mtoc,
            aircraft_data.mlw,
            aircraft_data.mzfw,
            flight_data.performance_tow,
            trimset_df,
            fuel_distribution=st.session_state.calculation_state.fuel_distribution,
            fuel_mode=st.session_state.calculation_state.fuel_mode,
            tail=aircraft_data.tail,
            ballast_fuel=st.session_state.get("computed_ballast_fuel", 0.0),
            performance_lw=flight_data.performance_lw
        )

        mtow_used = final_results["mtow_dynamic"] if tail != "N342AV" else aircraft_data.mtoc
        mtow_formula = final_results["mtow_formula"] if tail != "N342AV" else None
        mzfw_used = final_results["mzfw_dynamic"] if tail != "N342AV" else aircraft_data.mzfw
        mzfw_formula = final_results["mzfw_formula"] if tail != "N342AV" else None

        alerts = []
        if final_results["tow"] > mtow_used:
            alerts.append(f"TOW ({final_results['tow']:.1f} kg) excede el {'MTOWD' if tail != 'N342AV' else 'MTOW'} ({mtow_used:.1f} kg).")
        if performance_tow > 0 and final_results["tow"] > performance_tow:
            alerts.append(f"TOW ({final_results['tow']:.1f} kg) excede el Performance TOW ({performance_tow:.1f} kg).")
        if final_results["lw"] > aircraft_data.mlw:
            alerts.append(f"LW ({final_results['lw']:.1f} kg) excede el MLW ({aircraft_data.mlw:.1f} kg).")
        if performance_lw > 0 and final_results["lw"] > performance_lw:
            alerts.append(f"LW ({final_results['lw']:.1f} kg) excede el Performance LW ({performance_lw:.1f} kg).")
        if final_results["zfw_peso"] > mzfw_used:
            alerts.append(f"ZFW ({final_results['zfw_peso']:.1f} kg) excede el {'MZFWD' if tail != 'N342AV' else 'MZFW'} ({mzfw_used:.1f} kg).")

        complies, validation_df = check_cumulative_weights(df_asignados, cumulative_restrictions_fwd_df, cumulative_restrictions_aft_df)

        ldf_weight = df_asignados[df_asignados["Bodega"] == "LDF"]["Weight (KGS)"].sum() if not df_asignados[df_asignados["Bodega"] == "LDF"].empty else 0.0
        lda_weight = df_asignados[df_asignados["Bodega"] == "LDA"]["Weight (KGS)"].sum() if not df_asignados[df_asignados["Bodega"] == "LDA"].empty else 0.0
        ldf_complies = ldf_weight <= aircraft_data.ldf_limit
        lda_complies = lda_weight <= aircraft_data.lda_limit
        complies = complies and ldf_complies and lda_complies

        pallets_imbalance = 0.0
        if not df_asignados.empty:
            relevant_pallets = df_asignados[
                (df_asignados["Bodega"].isin(["MD", "LDA", "LDF"])) &
                (df_asignados["Y-arm"] != 0)
            ]
            left_weight = relevant_pallets[relevant_pallets["Y-arm"] < 0]["Weight (KGS)"].sum()
            right_weight = relevant_pallets[relevant_pallets["Y-arm"] > 0]["Weight (KGS)"].sum()
            pallets_imbalance = abs(left_weight - right_weight)

        st.markdown('<div id="validation_section"></div>', unsafe_allow_html=True)
        st.subheader("Validación de Pesos Acumulativos")
        with st.expander("Ver Validación de Pesos Acumulativos", expanded=False):
            st.dataframe(validation_df, use_container_width=True)

            if not complies:
                if "Posición Asignada" in validation_df.columns:
                    non_compliant_positions = validation_df[validation_df["Cumple"] == "No"]["Posición Asignada"].tolist()
                    if non_compliant_positions:
                        st.error(f"Las siguientes posiciones no cumplen los pesos acumulativos: {', '.join(non_compliant_positions)}")
                        alerts.append(f"Restricciones acumulativas no cumplidas en posiciones: {', '.join(non_compliant_positions)}")
                    else:
                        st.error("No se encontraron posiciones no conformes, pero el cálculo indica incumplimiento.")
                        alerts.append("Validación inconsistente: No hay posiciones no conformes pero complies=False.")
                else:
                    st.error("Error: La columna 'Posición Asignada' no está presente en los datos de validación.")
                    alerts.append("Error en la validación: Columna 'Posición Asignada' no encontrada.")
            else:
                st.success("Todas las posiciones cumplen los pesos acumulativos.")

        st.subheader("Validación de Límites de Peso por Bodega")
        if ldf_weight > aircraft_data.ldf_limit:
            st.error(f"El peso total en LDF ({ldf_weight:.1f} kg) excede el límite permitido ({aircraft_data.ldf_limit:.1f} kg).")
            alerts.append(f"Peso en LDF ({ldf_weight:.1f} kg) excede LDF_LIMIT ({aircraft_data.ldf_limit:.1f} kg).")
        else:
            st.success(f"El peso total en LDF ({ldf_weight:.1f} kg) está dentro del límite permitido ({aircraft_data.ldf_limit:.1f} kg).")

        if lda_weight > aircraft_data.lda_limit:
            st.error(f"El peso total en LDA ({lda_weight:.1f} kg) excede el límite permitido ({aircraft_data.lda_limit:.1f} kg).")
            alerts.append(f"Peso en LDA ({lda_weight:.1f} kg) excede LDA_LIMIT ({aircraft_data.lda_limit:.1f} kg).")
        else:
            st.success(f"El peso total en LDA ({lda_weight:.1f} kg) está dentro del límite permitido ({aircraft_data.lda_limit:.1f} kg).")





       # Load add_removal.csv to calculate add_removal_weight
        add_removal_path = os.path.join(aircraft_folder, "add_removal.csv")
        add_removal_weight = 0.0
        if os.path.exists(add_removal_path):
            try:
                add_removal_df = pd.read_csv(add_removal_path, sep=";", decimal=",")
                if "Weight" in add_removal_df.columns:
                    add_removal_weight = pd.to_numeric(add_removal_df["Weight"], errors="coerce").sum()
                else:
                    st.warning("El archivo add_removal.csv no contiene la columna 'Weight'.")
            except Exception as e:
                st.warning(f"Error al cargar add_removal.csv: {str(e)}")
        else:
            st.info("No se encontró add_removal.csv. Se asume peso removido/adicionado de 0 kg.")

        # Calculate adjusted_bow
        adjusted_bow = bow + add_removal_weight

        st.markdown('<div id="summary_section"></div>', unsafe_allow_html=True)
        st.subheader("Resumen Final de Peso y Balance")
        print_final_summary(
            df_asignados=df_asignados,
            operador=flight_data.operador,
            numero_vuelo=flight_data.numero_vuelo,
            matricula=flight_data.matricula,
            fecha_vuelo=flight_data.fecha_vuelo,
            hora_vuelo=flight_data.hora_vuelo,
            ruta_vuelo=flight_data.ruta_vuelo,
            revision=flight_data.revision,
            oew=aircraft_data.oew,
            bow=st.session_state.calculation_state.bow,
            add_removal_weight=add_removal_weight,
            adjusted_bow=adjusted_bow,
            peso_total=final_results.get("peso_total", 0.0),
            zfw_peso=final_results.get("zfw_peso", 0.0),
            zfw_mac=final_results.get("zfw_mac", 0.0),
            mzfw=mzfw_used,
            tow=final_results.get("tow", 0.0),
            tow_mac=final_results.get("tow_mac", 0.0),
            mtow=mtow_used,
            trip_fuel=flight_data.trip_fuel,
            lw=final_results.get("lw", 0.0),
            lw_mac=final_results.get("lw_mac", 0.0),
            mlw=aircraft_data.mlw,
            underload=final_results.get("underload", 0.0),
            mrow=final_results.get("mrow", 0.0),
            mrow_mac=final_results.get("mrow_mac", 0.0),
            takeoff_runway=flight_data.takeoff_runway,
            flaps_conf=flight_data.flaps_conf,
            temperature=flight_data.temperature,
            anti_ice=flight_data.anti_ice,
            air_condition=flight_data.air_condition,
            lateral_imbalance=final_results.get("lateral_imbalance", 0.0),
            max_payload_lw=final_results.get("max_payload_lw", 0.0),
            max_payload_tow=final_results.get("max_payload_tow", 0.0),
            max_payload_zfw=final_results.get("max_payload_zfw", 0.0),
            pitch_trim=final_results.get("pitch_trim", 0.0),
            complies=complies,
            validation_df=validation_df,
            fuel_table=fuel_table,
            fuel_tow=fuel_for_tow,
            fuel_lw=fuel_for_lw,
            mrw_limit=aircraft_data.mrw_limit,
            lateral_imbalance_limit=aircraft_data.lateral_imbalance_limit,
            fuel_distribution=st.session_state.calculation_state.fuel_distribution,
            fuel_mode=st.session_state.calculation_state.fuel_mode,
            ballast_fuel=st.session_state.get("computed_ballast_fuel", 0.0),
            performance_tow=flight_data.performance_tow,
            active_restrictions=active_restrictions,
            performance_lw=flight_data.performance_lw,
            ldf_weight=ldf_weight,
            ldf_limit=aircraft_data.ldf_limit,
            lda_weight=lda_weight,
            lda_limit=aircraft_data.lda_limit,
            mzfw_formula=mzfw_formula,
            mtow_formula=mtow_formula,
            qnh=flight_data.qnh,
            rwy_condition=flight_data.rwy_condition
        )

        # Display alerts if any
        if alerts:
            for alert in alerts:
                st.error(alert)

        # Envelope Section
        st.markdown('<div id="envelope_section"></div>', unsafe_allow_html=True)
        st.subheader("Envelope")
        plt_envelope = None
        temp_results = None
        try:
            if tail == "N342AV":
                try:
                    from N342AV_envelope import plot_cg_envelope
                except ImportError:
                    st.error("No se encontró N342AV_envelope.py")
                    return
            elif tail == "N337QT":
                try:
                    from N337QT_envelope import plot_cg_envelope
                except ImportError:
                    st.error("No se encontró N337QT_envelope.py")
                    return
            elif tail == "N338QT":
                try:
                    from N338QT_envelope import plot_cg_envelope
                except ImportError:
                    st.error("No se encontró N338QT_envelope.py")
                    return
            else:
                try:
                    from A330_200F_envelope import plot_cg_envelope
                except ImportError:
                    st.error("No se encontró A330_200F_envelope.py")
                    return

            temp_results = calculate_final_values(
                df_asignados if not df_asignados.empty else pd.DataFrame(columns=df.columns),
                st.session_state.calculation_state.bow,
                st.session_state.calculation_state.bow_moment_x,
                st.session_state.calculation_state.bow_moment_y,
                flight_data.fuel_kg,
                flight_data.taxi_fuel,
                flight_data.trip_fuel,
                st.session_state.calculation_state.moment_x_fuel_tow,
                st.session_state.calculation_state.moment_y_fuel_tow,
                st.session_state.calculation_state.moment_x_fuel_lw,
                st.session_state.calculation_state.moment_y_fuel_lw,
                aircraft_data.lemac,
                aircraft_data.mac_length,
                aircraft_data.mtoc,
                aircraft_data.mlw,
                aircraft_data.mzfw,
                flight_data.performance_tow,
                trimset_df,
                fuel_distribution=st.session_state.calculation_state.fuel_distribution,
                fuel_mode=st.session_state.calculation_state.fuel_mode,
                tail=aircraft_data.tail,
                ballast_fuel=st.session_state.get("computed_ballast_fuel", 0.0),
                performance_lw=flight_data.performance_lw
            )
            required_keys = ["zfw_peso", "zfw_mac", "tow", "tow_mac", "lw", "lw_mac"]
            missing_keys = [k for k in required_keys if k not in temp_results or temp_results[k] is None or np.isnan(temp_results[k])]
            if missing_keys:
                st.warning(f"No se puede graficar el envelope. Faltan o son inválidos: {', '.join(missing_keys)}")
            else:
                # Plot the envelope and get envelope data
                envelope_data = plot_cg_envelope(
                    temp_results["zfw_peso"],
                    temp_results["zfw_mac"],
                    temp_results["tow"],
                    temp_results["tow_mac"],
                    temp_results["lw"],
                    temp_results["lw_mac"]
                )
                plt_envelope = envelope_data["fig"]
                st.pyplot(plt_envelope)
                plt.close(plt_envelope)  # Close the figure to free memory

                # Function to interpolate %MAC limit at a given weight
                def interpolate_limit(weight, x_vals, y_vals):
                    if not x_vals or not y_vals:
                        return None
                    # Sort points by weight
                    sorted_points = sorted(zip(y_vals, x_vals), key=lambda x: x[0])
                    w = [p[0] for p in sorted_points]
                    m = [p[1] for p in sorted_points]
                    # Handle out-of-range cases
                    if weight < w[0]:
                        return m[0]
                    if weight > w[-1]:
                        return m[-1]
                    # Linear interpolation
                    for i in range(len(w) - 1):
                        if w[i] <= weight <= w[i + 1]:
                            fraction = (weight - w[i]) / (w[i + 1] - w[i])
                            return m[i] + fraction * (m[i + 1] - m[i])
                    return m[-1]

                # Validate CG values against envelope limits
                # ZFW CG (use cruise envelope, as ZFW is typically evaluated in cruise condition)
                zfw_weight = temp_results["zfw_peso"]
                zfw_mac_proj = envelope_data["projected_cg"]["zfw"]
                zfw_limit_fwd = interpolate_limit(zfw_weight, envelope_data["cruise"]["fwd"]["x"], envelope_data["cruise"]["fwd"]["y"])
                zfw_limit_aft = interpolate_limit(zfw_weight, envelope_data["cruise"]["aft"]["x"], envelope_data["cruise"]["aft"]["y"])
                if zfw_limit_fwd is not None and zfw_limit_aft is not None:
                    if zfw_mac_proj < zfw_limit_fwd:
                        alerts.append(f"ZFW CG ({zfw_mac_proj:.1f}% MAC) está más adelante del límite FWD ({zfw_limit_fwd:.1f}% MAC) en condición de crucero.")
                    elif zfw_mac_proj > zfw_limit_aft:
                        alerts.append(f"ZFW CG ({zfw_mac_proj:.1f}% MAC) está más atrás del límite AFT ({zfw_limit_aft:.1f}% MAC) en condición de crucero.")
                else:
                    alerts.append("No se pudieron validar los límites de ZFW CG debido a datos de envolvente inválidos.")

                # TOW CG (use takeoff envelope)
                tow_weight = temp_results["tow"]
                tow_mac_proj = envelope_data["projected_cg"]["tow"]
                tow_limit_fwd = interpolate_limit(tow_weight, envelope_data["takeoff"]["fwd"]["x"], envelope_data["takeoff"]["fwd"]["y"])
                tow_limit_aft = interpolate_limit(tow_weight, envelope_data["takeoff"]["aft"]["x"], envelope_data["takeoff"]["aft"]["y"])
                if tow_limit_fwd is not None and tow_limit_aft is not None:
                    if tow_mac_proj < tow_limit_fwd:
                        alerts.append(f"TOW CG ({tow_mac_proj:.1f}% MAC) está más adelante del límite FWD ({tow_limit_fwd:.1f}% MAC) en condición de despegue.")
                    elif tow_mac_proj > tow_limit_aft:
                        alerts.append(f"TOW CG ({tow_mac_proj:.1f}% MAC) está más atrás del límite AFT ({tow_limit_aft:.1f}% MAC) en condición de despegue.")
                else:
                    alerts.append("No se pudieron validar los límites de TOW CG debido a datos de envolvente inválidos.")

                # LW CG (use landing envelope)
                lw_weight = temp_results["lw"]
                lw_mac_proj = envelope_data["projected_cg"]["lw"]
                lw_limit_fwd = interpolate_limit(lw_weight, envelope_data["landing"]["fwd"]["x"], envelope_data["landing"]["fwd"]["y"])
                lw_limit_aft = interpolate_limit(lw_weight, envelope_data["landing"]["aft"]["x"], envelope_data["landing"]["aft"]["y"])
                if lw_limit_fwd is not None and lw_limit_aft is not None:
                    if lw_mac_proj < lw_limit_fwd:
                        alerts.append(f"LW CG ({lw_mac_proj:.1f}% MAC) está más adelante del límite FWD ({lw_limit_fwd:.1f}% MAC) en condición de aterrizaje.")
                    elif lw_mac_proj > lw_limit_aft:
                        alerts.append(f"LW CG ({lw_mac_proj:.1f}% MAC) está más atrás del límite AFT ({lw_limit_aft:.1f}% MAC) en condición de aterrizaje.")
                else:
                    alerts.append("No se pudieron validar los límites de LW CG debido a datos de envolvente inválidos.")

        except Exception as e:
            st.error(f"Error al generar el envelope: {str(e)}")

        # Distribution Section
        st.markdown('<div id="distribution_section"></div>', unsafe_allow_html=True)
        st.subheader("Distribución de Pallets")
        st.markdown('<div id="main_deck_distribution_section"></div>', unsafe_allow_html=True)

        # Function to send images and data to Flask server
        def send_images_to_flask(main_deck_base64, lower_decks_base64, total_carga, tow_cg, lateral_imbalance, pallets_imbalance, zfw_cg, lw_cg):
            try:
                response = requests.post(
                    "http://localhost:5000/update_images",
                    json={
                        "main_deck_base64": main_deck_base64,
                        "lower_decks_base64": lower_decks_base64,
                        "total_carga": total_carga,
                        "tow_cg": tow_cg,
                        "lateral_imbalance": lateral_imbalance,
                        "pallets_imbalance": pallets_imbalance,
                        "zfw_cg": zfw_cg,
                        "lw_cg": lw_cg
                    },
                    timeout=90
                )
                if response.status_code != 200:
                    st.warning("No se pudo enviar las imágenes al servidor Flask.")
            except requests.RequestException as e:
                st.warning(f"Error al enviar imágenes al servidor Flask: {str(e)}")

        # Generate plots with optimized size
        st.write("**Main Deck**")
        main_deck_fig = plot_main_deck(df_asignados, restricciones_df)
        if main_deck_fig:
            main_deck_fig.set_size_inches(18, 5)
            main_deck_img = BytesIO()
            main_deck_fig.savefig(main_deck_img, format="jpeg", bbox_inches="tight", dpi=100)
            main_deck_img.seek(0)
            st.session_state.main_deck_base64 = base64.b64encode(main_deck_img.getvalue()).decode('utf-8')
            st.pyplot(main_deck_fig)
            plt.close(main_deck_fig)
        else:
            st.session_state.main_deck_base64 = None
            st.warning("No se pudo generar la gráfica de Main Deck.")

        st.write("**Lower Decks**")
        lower_decks_fig = plot_lower_decks(df_asignados, restricciones_df)
        if lower_decks_fig:
            lower_decks_fig.set_size_inches(18, 5)
            lower_decks_img = BytesIO()
            lower_decks_fig.savefig(lower_decks_img, format="jpeg", bbox_inches="tight", dpi=100)
            lower_decks_img.seek(0)
            st.session_state.lower_decks_base64 = base64.b64encode(lower_decks_img.getvalue()).decode('utf-8')
            st.pyplot(lower_decks_fig)
            plt.close(lower_decks_fig)
        else:
            st.session_state.lower_decks_base64 = None
            st.warning("No se pudo generar la gráfica de Lower Decks.")

        # Send images and data to Flask server immediately the first time
        if (st.session_state.get('main_deck_base64') or st.session_state.get('lower_decks_base64')) and not st.session_state.get('images_sent', False):
            total_carga = df_asignados["Weight (KGS)"].sum() if not df_asignados.empty else 0.0
            send_images_to_flask(
                st.session_state.get('main_deck_base64'),
                st.session_state.get('lower_decks_base64'),
                total_carga,
                final_results.get("tow_mac", 0.0),
                final_results.get("lateral_imbalance", 0.0),
                pallets_imbalance,
                final_results.get("zfw_mac", 0.0),
                final_results.get("lw_mac", 0.0)
            )
            st.session_state.images_sent = True

        # Send images and data to Flask server every 3 seconds for updates
        if st.session_state.get('main_deck_base64') or st.session_state.get('lower_decks_base64'):
            if 'last_update_time' not in st.session_state:
                st.session_state.last_update_time = time.time()

            current_time = time.time()
            if current_time - st.session_state.last_update_time >= 3:
                total_carga = df_asignados["Weight (KGS)"].sum() if not df_asignados.empty else 0.0
                send_images_to_flask(
                    st.session_state.get('main_deck_base64'),
                    st.session_state.get('lower_decks_base64'),
                    total_carga,
                    final_results.get("tow_mac", 0.0),
                    final_results.get("lateral_imbalance", 0.0),
                    pallets_imbalance,
                    final_results.get("zfw_mac", 0.0),
                    final_results.get("lw_mac", 0.0)
                )
                st.session_state.last_update_time = current_time

            st.markdown(
                """
                <style>
                .custom-button {
                    background-color: #4CAF50;
                    color: white;
                    padding: 10px 20px;
                    border: none;
                    border-radius: 5px;
                    cursor: pointer;
                    font-size: 16px;
                    text-align: center;
                    display: inline-block;
                    text-decoration: none;
                }
                .custom-button:hover {
                    background-color: #45a009;
                }
                </style>
                <a href="http://localhost:5000/pallet_distribution" target="_blank" class="custom-button">Ver LIR Paralela</a>
                """,
                unsafe_allow_html=True
            )

        # Load Summary Data
        st.markdown('<div id="load_summary_data"></div>', unsafe_allow_html=True)
        st.markdown('<div class="summary-box">', unsafe_allow_html=True)
        st.markdown('<div class="summary-title">📊 Resumen de Carga</div>', unsafe_allow_html=True)
        col_weight_cg, col_imbalance_cg, col_cg_values = st.columns(3)
        with col_weight_cg:
            total_carga = df_asignados["Weight (KGS)"].sum() if not df_asignados.empty else 0.0
            st.markdown(f'<div class="summary-item"><b>Peso Total Carga Asignada:</b> {total_carga:,.1f} kg</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="summary-item"><b>TOW CG:</b> {final_results.get("tow_mac", 0.0):,.1f}% MAC</div>', unsafe_allow_html=True)
        with col_imbalance_cg:
            st.markdown(f'<div class="summary-item"><b>Desbalance Lateral:</b> {final_results.get("lateral_imbalance", 0.0):,.1f} kg.m (Límite: {aircraft_data.lateral_imbalance_limit:,.1f} kg.m)</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="summary-item"><b>Desbalance de Pallets (MD, LDA, LDF):</b> {pallets_imbalance:,.1f} kg</div>', unsafe_allow_html=True)
        with col_cg_values:
            st.markdown(f'<div class="summary-item"><b>ZFW CG:</b> {final_results.get("zfw_mac", 0.0):,.1f}% MAC</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="summary-item"><b>LW CG:</b> {final_results.get("lw_mac", 0.0):,.1f}% MAC</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        # Modal for Envelope
        if st.session_state.get("show_envelope", False):
            with st.container():
                st.markdown(
                    """
                    <style>
                    .modal {
                        position: fixed;
                        top: 50%;
                        left: 50%;
                        transform: translate(-50%, -50%);
                        background-color: white;
                        padding: 20px;
                        border-radius: 10px;
                        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
                        z-index: 1000;
                        max-width: 80%;
                        max-height: 80%;
                        overflow: auto;
                    }
                    </style>
                    <div class="modal">
                    """,
                    unsafe_allow_html=True
                )
                st.write("### Envelope")
                try:
                    plt.figure(figsize=(8, 6))
                    if tail == "N342AV":
                        from N342AV_envelope import plot_cg_envelope
                    elif tail == "N337QT":
                        from N337QT_envelope import plot_cg_envelope
                    elif tail == "N338QT":
                        from N338QT_envelope import plot_cg_envelope
                    else:
                        from A330_200F_envelope import plot_cg_envelope
                    plot_cg_envelope(
                        temp_results["zfw_peso"],
                        temp_results["zfw_mac"],
                        temp_results["tow"],
                        temp_results["tow_mac"],
                        temp_results["lw"],
                        temp_results["lw_mac"]
                    )
                    st.pyplot(plt.gcf())
                    plt.close()
                except Exception as e:
                    st.error(f"Error al generar el envelope: {str(e)}")
                if st.button("Cerrar", key="close_envelope"):
                    st.session_state.close_envelope_trigger = True
                st.markdown("</div>", unsafe_allow_html=True)

        if "close_envelope_trigger" in st.session_state and st.session_state.close_envelope_trigger:
            st.session_state.show_envelope = False
            del st.session_state.close_envelope_trigger
            st.rerun()

        # ... (rest of weight_balance.py code until the export section) ...

        # Export Section
        # Export Section
        st.markdown('<div id="export_section"></div>', unsafe_allow_html=True)
        st.subheader("Exportación")

        if alerts:
            st.warning("Hay alertas pendientes que podrían afectar la exportación:")
            for alert in alerts:
                st.write(f"- {alert}")

        if "calculation_state" in st.session_state and st.session_state.calculation_state.df is not None:
            data_to_save = {
                "flight_info": {
                    "operador": flight_data.operador,
                    "numero_vuelo": flight_data.numero_vuelo,
                    "matricula": flight_data.matricula,
                    "fecha_vuelo": flight_data.fecha_vuelo,
                    "hora_vuelo": flight_data.hora_vuelo,
                    "ruta_vuelo": flight_data.ruta_vuelo,
                    "revision": flight_data.revision,
                    "destino_inicial": flight_data.destino_inicial,
                },
                "calculated_values": {
                    "oew": aircraft_data.oew,
                    "bow": st.session_state.calculation_state.bow,
                    "add_removal_weight": add_removal_weight,
                    "adjusted_bow": adjusted_bow,
                    "fuel_kg": flight_data.fuel_kg,
                    "trip_fuel": flight_data.trip_fuel,
                    "taxi_fuel": flight_data.taxi_fuel,
                    "ballast_fuel": st.session_state.get("computed_ballast_fuel", 0.0),
                    "moment_x_fuel_tow": st.session_state.calculation_state.moment_x_fuel_tow,
                    "moment_y_fuel_tow": st.session_state.calculation_state.moment_y_fuel_tow,
                    "moment_x_fuel_lw": st.session_state.calculation_state.moment_x_fuel_lw,
                    "moment_y_fuel_lw": st.session_state.calculation_state.moment_y_fuel_lw,
                    "zfw_peso": final_results.get("zfw_peso", 0.0),
                    "zfw_mac": final_results.get("zfw_mac", 0.0),
                    "tow": final_results.get("tow", 0.0),
                    "tow_mac": final_results.get("tow_mac", 0.0),
                    "mrow": final_results.get("mrow", 0.0),
                    "mrow_mac": final_results.get("mrow_mac", 0.0),
                    "lw": final_results.get("lw", 0.0),
                    "lw_mac": final_results.get("lw_mac", 0.0),
                    "underload": final_results.get("underload", 0.0),
                    "max_payload_lw": final_results.get("max_payload_lw", 0.0),
                    "max_payload_tow": final_results.get("max_payload_tow", 0.0),
                    "max_payload_zfw": final_results.get("max_payload_zfw", 0.0),
                    "mzfw_dynamic": mzfw_used,
                    "mzfw_formula": mzfw_formula,
                    "mtow_dynamic": mtow_used,
                    "mtow_formula": mtow_formula,
                    "fuel_distribution": st.session_state.calculation_state.fuel_distribution,
                    "fuel_mode": st.session_state.calculation_state.fuel_mode
                },
                "passengers": {
                    "cockpit": flight_data.passengers_cockpit,
                    "supernumerary": flight_data.passengers_supernumerary,
                    "cockpit_weight": st.session_state.calculation_state.passengers_cockpit_total_weight,
                    "cockpit_moment_x": st.session_state.calculation_state.passengers_cockpit_total_moment_x,
                    "supernumerary_weight": st.session_state.calculation_state.passengers_supernumerary_total_weight,
                    "supernumerary_moment_x": st.session_state.calculation_state.passengers_supernumerary_total_moment_x,
                },
                "takeoff_conditions": {
                    "runway": flight_data.takeoff_runway,
                    "rwy_condition": flight_data.rwy_condition,
                    "flaps_conf": flight_data.flaps_conf,
                    "temperature": flight_data.temperature,
                    "air_condition": flight_data.air_condition,
                    "anti_ice": flight_data.anti_ice,
                    "qnh": flight_data.qnh,
                    "performance_tow": flight_data.performance_tow,
                    "performance_lw": flight_data.performance_lw,
                },
                "manifest_data": st.session_state.calculation_state.df.to_dict(orient="records") if st.session_state.calculation_state.df is not None else [],
                "posiciones_usadas": list(st.session_state.calculation_state.posiciones_usadas),
                "rotaciones": st.session_state.calculation_state.rotaciones,
                "tipo_carga": flight_data.tipo_carga,
            }
        else:
            st.error("No hay datos de cálculo para exportar. Por favor, complete un cálculo primero.")
            data_to_save = None

        # Define output folder
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(script_dir, "Output")
        os.makedirs(output_dir, exist_ok=True)

        # Get user information
        full_name = st.session_state.get("full_name", "UsuarioDesconocido")
        user_license = "SinLicencia"
        users_json_path = os.path.join(script_dir, "users.json")
        try:
            if os.path.exists(users_json_path):
                with open(users_json_path, "r", encoding="utf-8") as f:
                    users = json.load(f)
                for user in users:
                    if user.get("Usuario") == st.session_state.get("username"):
                        user_license = user.get("Licencia", "SinLicencia")
                        break
        except Exception as e:
            st.warning(f"Error al leer users.json: {str(e)}")

        # Sanitize full_name and user_license for filename
        def sanitize_filename(s):
            return "".join(c for c in s if c.isalnum() or c in ('_', '-')).replace(" ", "_")

        sanitized_full_name = sanitize_filename(full_name)
        sanitized_license = sanitize_filename(user_license)

        # Define base filenames
        execution_time = datetime.now().strftime("%H%M%S")
        fecha_vuelo_safe = flight_data.fecha_vuelo.replace("/", "_")
        json_base_name = f"{aircraft_data.tail}_{flight_data.numero_vuelo}_{flight_data.ruta_vuelo}_{fecha_vuelo_safe}_{flight_data.revision}_W&B_{sanitized_full_name}_{sanitized_license}"
        excel_base_name = f"{aircraft_data.tail}_{flight_data.numero_vuelo}_{flight_data.ruta_vuelo}_{fecha_vuelo_safe}_{flight_data.revision}_W&B"
        json_path = os.path.join(output_dir, f"{json_base_name}.json")
        excel_path = os.path.join(output_dir, f"{excel_base_name}.xlsm")

        # Get unique filenames
        json_save_path = get_unique_filename(json_path, "json")
        excel_save_path = get_unique_filename(excel_path, "xlsm")

        if st.button("Exportar Documentos", key="export_documents"):
            if data_to_save is None:
                st.error("No se pudieron generar los documentos debido a datos insuficientes.")
            else:
                try:
                    # Export JSON
                    json_str = json.dumps(data_to_save, indent=4, ensure_ascii=False, cls=NumpyEncoder)
                    json_bytes = json_str.encode('utf-8')
                    with open(json_save_path, "wb") as f:
                        f.write(json_bytes)
                    json_buffer = BytesIO(json_bytes)

                    # Export Excel
                    template_path = os.path.join(script_dir, "templates", "template.xlsm")
                    if not os.path.exists(template_path):
                        st.error(f"No se encontró el archivo de plantilla en: {template_path}")
                        return

                    wb = load_workbook(template_path, keep_vba=True)
                    ws = wb.active

                    # Fill in the cells
                    ws['A3'] = flight_data.operador
                    ws['C3'] = flight_data.numero_vuelo
                    ws['D3'] = flight_data.fecha_vuelo
                    ws['E3'] = flight_data.matricula
                    ws['F3'] = flight_data.ruta_vuelo
                    ws['H3'] = flight_data.revision
                    ws['B4'] = aircraft_data.oew
                    ws['B5'] = st.session_state.get("computed_ballast_fuel", 0.0)
                    ws['B6'] = add_removal_weight
                    ws['B7'] = adjusted_bow
                    ws['B8'] = final_results.get("peso_total", 0.0)
                    ws['B9'] = mzfw_used
                    ws['B10'] = fuel_for_tow
                    ws['B12'] = flight_data.trip_fuel
                    ws['B13'] = flight_data.taxi_fuel
                    ws['B15'] = final_results.get("underload", 0.0)
                    ws['B16'] = final_results.get("mrow", 0.0)
                    ws['B18'] = final_results.get("zfw_peso", 0.0)
                    ws['B19'] = final_results.get("tow", 0.0)
                    ws['B20'] = final_results.get("lw", 0.0)
                    ws['B21'] = final_results.get("pitch_trim", 0.0)
                    ws['B23'] = flight_data.passengers_cockpit + 2
                    ws['B24'] = flight_data.passengers_supernumerary
                    ws['B36'] = mtow_used
                    ws['B37'] = flight_data.performance_tow
                    ws['B40'] = aircraft_data.mlw
                    ws['B41'] = flight_data.performance_lw
                    ws['C18'] = final_results.get("zfw_mac", 0.0)
                    ws['C19'] = final_results.get("tow_mac", 0.0)
                    ws['C20'] = final_results.get("lw_mac", 0.0)

                    bodega_summary = df_asignados.groupby("Bodega")["Weight (KGS)"].sum().reset_index()
                    ws['B25'] = bodega_summary[bodega_summary["Bodega"] == "MD"]["Weight (KGS)"].iloc[0] if not bodega_summary[bodega_summary["Bodega"] == "MD"].empty else 0.0
                    ws['B26'] = bodega_summary[bodega_summary["Bodega"] == "LDF"]["Weight (KGS)"].iloc[0] if not bodega_summary[bodega_summary["Bodega"] == "LDF"].empty else 0.0
                    ws['B27'] = bodega_summary[bodega_summary["Bodega"] == "LDA"]["Weight (KGS)"].iloc[0] if not bodega_summary[bodega_summary["Bodega"] == "LDA"].empty else 0.0
                    ws['B28'] = bodega_summary[bodega_summary["Bodega"] == "BULK"]["Weight (KGS)"].iloc[0] if not bodega_summary[bodega_summary["Bodega"] == "BULK"].empty else 0.0
                    ws['B29'] = final_results.get("lateral_imbalance", 0.0)
                    ws['E7'] = flight_data.takeoff_runway
                    ws['E8'] = flight_data.flaps_conf
                    ws['E9'] = flight_data.anti_ice
                    ws['E10'] = flight_data.air_condition
                    ws['E12'] = flight_data.temperature
                    ws['E13'] = flight_data.qnh

                    # Add User Full Name and License to cells
                    user_info = f"{full_name} - {user_license}"
                    ws['C31'] = user_info
                    ws['C32'] = user_info
                    ws['N31'] = user_info
                    ws['N32'] = user_info

                    # Insert images
                    if plt_envelope:
                        envelope_img = BytesIO()
                        plt_envelope.savefig(envelope_img, format="png", bbox_inches="tight", dpi=100)
                        envelope_img.seek(0)
                        img = OpenpyxlImage(envelope_img)
                        img.width = 400
                        img.height = 400
                        ws.add_image(img, 'G6')
                        plt.close(plt_envelope)

                    if st.session_state.get('main_deck_base64'):
                        main_deck_img = BytesIO(base64.b64decode(st.session_state.main_deck_base64))
                        img = OpenpyxlImage(main_deck_img)
                        img.width = 800
                        img.height = 200
                        ws.add_image(img, 'M6')

                    if st.session_state.get('lower_decks_base64'):
                        lower_decks_img = BytesIO(base64.b64decode(st.session_state.lower_decks_base64))
                        img = OpenpyxlImage(lower_decks_img)
                        img.width = 800
                        img.height = 200
                        ws.add_image(img, 'M21')

                    # Save Excel
                    wb.save(excel_save_path)
                    with open(excel_save_path, "rb") as f:
                        excel_buffer = BytesIO(f.read())
                    excel_buffer.seek(0)

                    # Provide downloads
                    col_dl1, col_dl2 = st.columns(2)
                    with col_dl1:
                        st.download_button(
                            label="Descargar JSON",
                            data=json_buffer,
                            file_name=os.path.basename(json_save_path),
                            mime="application/json"
                        )
                    with col_dl2:
                        st.download_button(
                            label="Descargar Excel",
                            data=excel_buffer,
                            file_name=os.path.basename(excel_save_path),
                            mime="application/vnd.ms-excel.sheet.macroEnabled.12"
                        )

                    st.success(f"Documentos generados.")
                except Exception as e:
                    st.error(f"Error al generar los documentos: {str(e)}")
        else:
            st.warning("Presiona boton para exportar.")

